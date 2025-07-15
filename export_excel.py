import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from io import BytesIO
from database import create_connection, database


def generar_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "SEGUIMIENTO DE ACTAS"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") 
    header_font = Font(color="FFFFFF", bold=True)
    
    # Estilos para filas de datos (para rayas de cebra)
    odd_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Blanco

    # Estilo de borde
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # --- Escribir Encabezados y Aplicar Estilos ---
    headers = ["CEDULA", "NOMBRE COMPLETO", "ZONA", "FORMATO", "FECHA", "OBSERVACION", "ESTADO"]
    for col_idx, header_text in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Ajustar el ancho de las columnas (opcional, pero mejora la presentación)
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 15

    # --- Conectar a la Base de Datos y Escribir Datos ---
    conn = create_connection(database) # Asumiendo que create_connection y database están definidos
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM actas")
            rows = cursor.fetchall()

            for row_idx, row_data in enumerate(rows, start=2): # Empieza desde la fila 2 para los datos
                # Determinar el estilo de fondo de la fila (rayas de cebra)
                current_row_fill = odd_row_fill if (row_idx % 2 != 0) else even_row_fill

                # Iterar sobre los datos de la fila y aplicarlos a las celdas
                # Asegúrate de que los índices de row_data coincidan con el orden de tus columnas en la DB
                # CEDULA (0), NOMBRE COMPLETO (1), ZONA (2), FORMATO (3), FECHA (4), OBSERVACION (5), ESTADO (6)
                data_to_write = [
                    row_data[1], # CEDULA
                    row_data[2], # NOMBRE COMPLETO
                    row_data[3], # ZONA
                    row_data[4], # FORMATO
                    row_data[5], # FECHA
                    row_data[6], # OBSERVACION
                    row_data[7]  # ESTADO
                ]

                for col_idx, value in enumerate(data_to_write, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = current_row_fill
                    cell.border = thin_border

        except sqlite3.Error as e:
            print(f"Error al leer datos de la base de datos: {e}")
            return None
        finally:
            if conn:
                conn.close()
    else:
        print("No se pudo conectar a la base de datos para obtener datos.")
        return None


    excel_file_buffer = BytesIO()
    workbook.save(excel_file_buffer)
    excel_file_buffer.seek(0)
    return excel_file_buffer
